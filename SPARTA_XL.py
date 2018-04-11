# coding=utf-8
#COM wrapper for Python
import traceback
import pandas as pd
#Oracle wrapper for Python
import cx_Oracle
#datetime for formatting strings and time for claculating execution time
import datetime,time
#To strip out the HTML/CSS from HP ALM description
from bs4 import BeautifulSoup
#To exit with zero and to modify the PATH env variable
import os,sys
#To register the OTA dll from Python
from ctypes import windll
#Excel module for Python
from openpyxl import Workbook
import sqlparse
import pyodbc
import re
import Tkinter as tk
from Tkinter import *
import ttk
from tkFileDialog import askopenfilename
import pickle
import os.path
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from multiprocessing import Process, Queue
#import pythoncom
import multiprocessing
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import string
import win32com.client
import sqlparse as MySQLdb
from datetime import datetime
# import sqlalchemy

STD_FONT=('Arial',10,'bold')
STD_HEAD=('Arial',12,'bold')

if os.path.exists('data.pickle') and os.stat("data.pickle").st_size != 0 :
    with open('data.pickle','rb') as handle:
		profile=pickle.load(handle)
		print profile
else:
	print '||NOTE||\n\t\t*********** No saved profiles!!! *************'
	profile = {'Profile':['1','1','1','1','1','1','1']}

class Index:
	def __init__(self,master):
		self.master=master
		# master.configure(background='black')
		# self.pic=PhotoImage(file='logo.gif')
		# self.robot=PhotoImage(file='robot.gif')
		self.frame=ttk.Frame(self.master)
		self.filllabel=tk.Label(self.master,text='',width=40,height=5)
		# self.logolabel=tk.Label(self.master,image=self.pic)
		# self.sparta=tk.Label(self.master,image=self.robot)
		self.logolabel=tk.Label(self.master,text='')
		self.sparta=tk.Label(self.master,text='')
		self.label=ttk.Label(self.master,text="Select Profile",font=STD_FONT)
		self.var=StringVar()
		self.warn=StringVar()
		self.var.set('Profile')
		self.warn.set('')
		self.option=ttk.OptionMenu(self.master,self.var,*profile,command = self.startNextSQLServer)
		self.button2=ttk.Button(self.master,text="Create Profile", command=self.createProfileSQLServer)	
		self.label2 = ttk.Label(self.master,textvariable=self.warn,foreground='red')
		self.filllabel.grid(row=1,column=1)
		self.sparta.grid(row=1,column=3,rowspan=8)
		self.logolabel.grid(row=3,column=2)
		self.label.grid(row=4,column=2)
		self.option.grid(row=5,column=2)
		self.button2.grid(row=6,column=2)
		self.label2.grid(row=7,column=2)
		self.frame.grid(row=8,column=2)

	def startNextSQLServer(self,val):
		self.val=val
		self.warn.set('')
		if self.val in profile:
			self.profileDetails=""
			qcuser= profile[self.val][0]
			qcpass= profile[self.val][1]
			SQLserver= profile[self.val][2]
			prof = profile[self.val]
		self.newWindow=tk.Toplevel(self.master)
		self.newWindow.title('Profile - {}'.format(self.val))
		self.newWindow.wm_geometry("600x280+350+150")
		self.newWindow.resizable(0,0)
		self.page=HomepageSQLServer(self.newWindow,prof)
	

	def createProfileSQLServer(self):
		self.newWindow2 = tk.Toplevel(self.master)
		self.newWindow2.title('Create Profile')
		self.newWindow2.wm_geometry("245x150+350+150")
		self.newWindow2.resizable(0,0)
		self.qcuname = StringVar()
		self.qcpass = StringVar()
		self.SQLserver = StringVar()
		self.profileName = StringVar()
		self.frame=ttk.Frame(self.newWindow2)
		self.label1 = ttk.Label(self.newWindow2,text="Enter QC UserName")
		self.entry1 = ttk.Entry(self.newWindow2,textvariable = self.qcuname)
		self.label2 = ttk.Label(self.newWindow2,text="Enter QC Password")
		self.entry2 = ttk.Entry(self.newWindow2,textvariable = self.qcpass, show='*')
		self.label3 = ttk.Label(self.newWindow2,text="Enter Server")
		self.entry3 = ttk.Entry(self.newWindow2,textvariable = self.SQLserver)
		self.label8 = ttk.Label(self.newWindow2,text="Enter Profile Name")
		self.entry8 = ttk.Entry(self.newWindow2,textvariable = self.profileName)
		self.button = ttk.Button(self.newWindow2,text='Save Profile', command = self.dispSQLServer)
		self.label1.grid(row=1,column=1)
		self.entry1.grid(row=1,column=2)
		self.label2.grid(row=2,column=1)
		self.entry2.grid(row=2,column=2)
		self.label3.grid(row=3,column=1)
		self.entry3.grid(row=3,column=2)
		self.label8.grid(row=4,column=1)
		self.entry8.grid(row=4,column=2)
		self.button.grid(row=5,column=2)

	

	def dispSQLServer(self):
		if self.qcuname.get() == "" or self.qcpass.get() == "" or self.SQLserver.get() == "" or self.profileName == "":
			self.warn.set('Form Incomplete. Profile not saved')
			self.newWindow2.destroy()
		else:
			self.warn.set('')
			self.value=list()
			self.value.append(self.qcuname.get())
			self.value.append(self.qcpass.get())
			self.value.append(self.SQLserver.get())
			self.key=self.profileName.get()
			profile[self.key]=self.value
			self.option.set_menu(*profile)
			with open('data.pickle', 'wb') as handle:
				pickle.dump(profile, handle, protocol=pickle.HIGHEST_PROTOCOL)
			self.newWindow2.destroy()

	

class HomepageSQLServer:

	def __init__(self,master,prof):
		self.master=master	
		self.prof=prof
		for key in prof:
			self.qcuser= prof[0]
			self.qcpass= prof[1]
			self.SQLserver= prof[2]
		print self.qcuser,self.qcpass,self.SQLserver	
		self.srcpath=StringVar()
		self.despath=StringVar()
		self.specpath=StringVar()
		self.testsetpath=StringVar()
		self.testsetname=StringVar()
		self.ddlpath=StringVar()
		self.qexesrcpath=StringVar()
		self.msg=StringVar()
		self.msg1=StringVar()
		self.spartamsg=StringVar()
		self.msg2=StringVar()
		self.msg3=StringVar()
		self.defectflag=IntVar()
		self.continuerunflag = IntVar()		
		self.msg.set('')
		self.msg1.set('')
		self.spartamsg.set('')
		self.msg2.set('')
		self.msg3.set('')
		self.qexetcname = ''
		# self.photo=PhotoImage(file="sea.gif")
		# self.secondary_process
		# self.q = Queue()
		# self.secondary_process =Process(target=call_Sparta,args=(self.defectflag.get(),self.q,self.testsetpath.get(),self.testsetname.get(),self.qcuser,self.qcpass,self.dbuser,self.dbpass,self.dbhost,self.dbport,self.dbserv))
		self.f1 = ttk.Frame(self.master)
		self.f2 = ttk.Frame(self.master)
		self.f3 = ttk.Frame(self.master)
		self.b1 = ttk.Button(self.f2,text='PyRite',command=self.frame1)
		self.b2 = ttk.Button(self.f2,text='SPEQ',command=self.frame2)
		self.b3 = ttk.Button(self.f2,text='ALM Executor',command=self.frame3)
		self.b4 = ttk.Button(self.f2,text= 'DDL Validator',command=self.frame4)
		self.b5 = ttk.Button(self.f2,text= 'Query Executor',command=self.frame5)
		self.l1 = ttk.Label(self.f3,text=' ')

		self.b1.grid(row=1,column=1,ipadx='10',ipady='3',padx='25')
		# self.b2.grid(row=1,column=2,ipadx='10',ipady='3',padx='25')
		self.b3.grid(row=1,column=3,ipadx='10',ipady='3',padx='25')
		# self.b4.grid(row=1,column=4,ipadx='10',ipady='3',padx='25')
		self.b5.grid(row=1,column=5,ipadx='10',ipady='3',padx='25')
		self.labelfill1=ttk.Label(self.f1,text=' ')
		self.labelfill2=ttk.Label(self.f1,text=' ')
		self.labelfill3=ttk.Label(self.f1,text=' ')
		self.labelfill4=ttk.Label(self.f1,text=' ')
		self.labelfill5=ttk.Label(self.f1,text=' ')
		self.labelfill6=ttk.Label(self.f1,text=' ')
		self.labelfill7=ttk.Label(self.f1,text=' ')
		self.labelfill8=ttk.Label(self.f1,text=' ')		# self.labelfill=ttk.Label(self.f1,text=' ')
		self.fill1 = ttk.Label(self.f1,text=' ')
		self.fill2 = ttk.Label(self.f1,text=' ')
		self.fill3 = ttk.Label(self.f1,text=' ')
		self.fill4 = ttk.Label(self.f1,text=' ')
		self.fill5 = ttk.Label(self.f1,text=' ')
		self.fill6 = ttk.Label(self.f1,text=' ')
		self.fill7 = ttk.Label(self.f1,text=' ')
		self.fill8 = ttk.Label(self.f1,text=' ')
		self.fill9 = ttk.Label(self.f1,text=' ')
		self.fill10 = ttk.Label(self.f1,text=' ')
		self.fill11 = ttk.Label(self.f1,text=' ')
		self.fill12 = ttk.Label(self.f1,text=' ')
		self.fill13 = ttk.Label(self.f1,text=' ')
		self.fill14 = ttk.Label(self.f1,text=' ')
		self.fill15 = ttk.Label(self.f1,text=' ')
		self.fill16 = ttk.Label(self.f1,text=' ')
		self.fill17 = ttk.Label(self.f1,text=' ')
		self.fill18 = ttk.Label(self.f1,text=' ')
		# self.fill11 = ttk.Label(self.f1,text=' ')
		self.label1=ttk.Label(self.f1,text='Enter Source File Path')
		self.entry1=ttk.Entry(self.f1,textvariable=self.srcpath,width=30,state=tk.DISABLED)
		# self.searchbutton1=ttk.Button(self.f1,image=self.photo, command=self.getsrc)
		self.searchbutton1=ttk.Button(self.f1,text='Search', command=self.getsrc)
		self.button1=ttk.Button(self.f1,text='Run Test Design - PyRite', command=self.call_PyRite)
		self.label10=ttk.Label(self.f1,text='Automatic English Test Case Creation',font=STD_HEAD,foreground='grey',background='dark blue')
		self.label3=ttk.Label(self.f1,textvariable=self.msg,foreground='limegreen')
		self.label7=ttk.Label(self.f1,text='Enter Mapping Spec Path')
		self.entry7=ttk.Entry(self.f1,textvariable=self.specpath,width=30,state=tk.DISABLED)
		# self.searchbutton2=ttk.Button(self.f1,image=self.photo, command=self.getsrc2)
		self.searchbutton2=ttk.Button(self.f1,text='Search', command=self.getsrc2)
		self.button2=ttk.Button(self.f1,text=' Run Test Design - SPEQ', command=self.call_SPEQ)
		self.label11=ttk.Label(self.f1,text='Automatic Query Generation for Data Validation',font=STD_HEAD,foreground='grey',background='dark blue')
		self.label8=ttk.Label(self.f1,textvariable=self.msg1,foreground='limegreen')
		self.label4=ttk.Label(self.f1,text='Enter Test Set Path')
		self.entry3=ttk.Entry(self.f1,textvariable=self.testsetpath,width=30)
		self.label5=ttk.Label(self.f1,text='Enter Test Set Name')
		self.entry4=ttk.Entry(self.f1,textvariable=self.testsetname,width=30)
		self.cbutton1=ttk.Checkbutton(self.f1,text='Enable Defect Creation',variable =self.defectflag)
		self.cbutton2=ttk.Checkbutton(self.f1,text='Continue Previous Run',variable =self.continuerunflag)
		self.button3=ttk.Button(self.f1,text='Run Test Execution', command=self.start_thread)
		self.label12=ttk.Label(self.f1,text='Automated Execution of Test Cases in ALM',font=STD_HEAD,foreground='grey',background='dark blue')
		self.label6=ttk.Label(self.f1,textvariable=self.spartamsg)
		self.label13=ttk.Label(self.f1,text='Automatic DDL Validation',font=STD_HEAD,foreground='grey',background='dark blue')
		self.label14=ttk.Label(self.f1,text='Enter DDL IPath')
		self.entry9=ttk.Entry(self.f1,textvariable=self.ddlpath,width=30,state=tk.DISABLED)
		# self.searchbutton3=ttk.Button(self.f1,image=self.photo, command=self.getsrc3)
		self.searchbutton3=ttk.Button(self.f1,text='Search', command=self.getsrc3)
		self.label15=ttk.Label(self.f1,textvariable=self.msg2,foreground='limegreen')
		self.button5=ttk.Button(self.f1,text=' Run DDL Validation', command=self.call_DDL)

		self.progbar = ttk.Progressbar(self.f1)
		self.progbar.config(maximum=10,mode='indeterminate',length=100)
		self.label16=ttk.Label(self.f1,text='Enter Source File Path')
		self.entry10=ttk.Entry(self.f1,textvariable=self.qexesrcpath,width=30,state=tk.DISABLED)
		# self.searchbutton4=ttk.Button(self.f1,image=self.photo, command=self.getsrc4)
		self.searchbutton4=ttk.Button(self.f1,text='Search', command=self.getsrc4)
		self.button6=ttk.Button(self.f1,text='Run Query Executor', command=self.start_qexethread)
		self.label17=ttk.Label(self.f1,text='Automated Execution of Test Cases from Excel',font=STD_HEAD,foreground='grey',background='dark blue')
		self.label18=ttk.Label(self.f1,textvariable=self.msg3,foreground='limegreen')
		self.button4=ttk.Button(self.f3,text='Close', command=self.closePage)

		self.labelfill1.grid(row=1,column=0)
		self.labelfill2.grid(row=2,column=0)
		self.labelfill3.grid(row=3,column=0)
		self.labelfill4.grid(row=4,column=0)
		self.labelfill5.grid(row=5,column=0)
		self.labelfill6.grid(row=6,column=0)
		self.labelfill7.grid(row=7,column=0)
		self.labelfill8.grid(row=8,column=0)

		self.label1.grid(row=4,column=0)
		self.entry1.grid(row=4,column=1)
		self.searchbutton1.grid(row=4,column=2)
		self.button1.grid(row=5,column=1)
		self.label10.grid(row=2,column=0,columnspan=3)
		self.label3.grid(row=6,column=0,columnspan=3)
		self.fill1.grid(row=1,column=0)
		self.fill2.grid(row=3,column=0)
		self.fill3.grid(row=7,column=0)
		self.fill4.grid(row=8,column=0)
		self.label1.grid_remove()
		self.entry1.grid_remove()
		self.searchbutton1.grid_remove()
		self.button1.grid_remove()
		self.label3.grid_remove()
		self.label10.grid_remove()
		self.fill1.grid_remove()
		self.fill2.grid_remove()
		self.fill3.grid_remove()
		self.fill4.grid_remove()
		self.label7.grid(row=4,column=0)
		self.entry7.grid(row=4,column=1)
		self.searchbutton2.grid(row=4,column=2)
		self.button2.grid(row=5,column=1)
		self.label11.grid(row=2,column=0,columnspan=3)
		self.label8.grid(row=6,column=0,columnspan=3)
		self.fill5.grid(row=1,column=0)
		self.fill6.grid(row=3,column=0)
		self.fill7.grid(row=7,column=0)
		self.fill8.grid(row=8,column=0)
		self.label7.grid_remove()
		self.entry7.grid_remove()
		self.searchbutton2.grid_remove()
		self.label8.grid_remove()
		self.button2.grid_remove()
		self.label11.grid_remove()
		self.fill5.grid_remove()
		self.fill6.grid_remove()
		self.fill7.grid_remove()
		self.fill8.grid_remove()
		self.label4.grid(row=4,column=0)
		self.entry3.grid(row=4,column=1)
		self.label5.grid(row=5,column=0)
		self.entry4.grid(row=5,column=1)
		self.cbutton1.grid(row=6,column=0)
		self.cbutton2.grid(row=6,column=1)
		self.button3.grid(row=7,column=1)
		self.label12.grid(row=2,column=0,columnspan=3)
		self.label6.grid(row=8,column=0,columnspan=3)
		self.progbar.grid(row=9,column=0,columnspan=3)
		self.fill9.grid(row=1,column=0)
		self.fill10.grid(row=3,column=0)
		# self.fill11.grid(row=3,column=0)
		self.label4.grid_remove()
		self.entry3.grid_remove()
		self.label5.grid_remove()
		self.entry4.grid_remove()
		self.cbutton1.grid_remove()
		self.cbutton2.grid_remove()
		self.button3.grid_remove()
		self.label12.grid_remove()		
		self.label6.grid_remove()
		self.progbar.grid_remove()
		self.fill9.grid_remove()
		self.fill10.grid_remove()
		self.fill11.grid(row=1,column=0)
		self.label13.grid(row=2,column=0,columnspan=3)
		self.fill12.grid(row=3,column=0)
		self.label14.grid(row=4,column=0)
		self.entry9.grid(row=4,column=1)
		self.searchbutton3.grid(row=4,column=2)
		self.button5.grid(row=5,column=1)
		self.label15.grid(row=6,column=0,columnspan=3)
		self.fill13.grid(row=7,column=0)
		self.fill14.grid(row=8,column=0)
		self.fill11.grid_remove()
		self.label13.grid_remove()
		self.fill12.grid_remove()
		self.label14.grid_remove()
		self.entry9.grid_remove()
		self.searchbutton3.grid_remove()
		self.button5.grid_remove()
		self.label15.grid_remove()
		self.fill13.grid_remove()
		self.fill14.grid_remove()
		self.label16.grid(row=4,column=0)
		self.entry10.grid(row=4,column=1)
		self.searchbutton4.grid(row=4,column=2)
		self.button6.grid(row=5,column=1)
		self.label17.grid(row=2,column=0,columnspan=3)
		self.label8.grid(row=6,column=0,columnspan=3)
		self.fill15.grid(row=1,column=0)
		self.fill16.grid(row=3,column=0)
		self.fill17.grid(row=7,column=0)
		self.fill18.grid(row=8,column=0)
		self.label16.grid_remove()
		self.entry10.grid_remove()
		self.searchbutton4.grid_remove()
		self.button6.grid_remove()
		self.label17.grid_remove()
		self.label18.grid_remove()
		self.fill15.grid_remove()
		self.fill16.grid_remove()
		self.fill17.grid_remove()
		self.fill18.grid_remove()


		self.l1.grid(row=1,column=1)
		self.button4.grid(row=2,column=1,ipadx='10',padx='15',ipady='3')

		self.f1.pack()
		self.f2.pack()
		self.f3.pack()

	def frame1(self):
		self.label1.grid()
		self.entry1.grid()
		self.searchbutton1.grid()
		self.button1.grid()
		self.label10.grid()
		self.label3.grid()
		self.fill1.grid()
		self.fill2.grid()
		self.fill3.grid()
		self.fill4.grid()
		self.msg.set('')
		self.srcpath.set('')

		self.labelfill1.grid_remove()
		self.labelfill2.grid_remove()
		self.labelfill3.grid_remove()
		self.labelfill4.grid_remove()
		self.labelfill5.grid_remove()
		self.labelfill6.grid_remove()
		self.labelfill7.grid_remove()
		self.labelfill8.grid_remove()
		self.label7.grid_remove()
		self.entry7.grid_remove()
		self.searchbutton2.grid_remove()
		self.label8.grid_remove()
		self.button2.grid_remove()
		self.fill5.grid_remove()
		self.fill6.grid_remove()
		self.fill7.grid_remove()
		self.fill8.grid_remove()
		self.label4.grid_remove()
		self.label11.grid_remove()
		self.label12.grid_remove()
		self.entry3.grid_remove()
		self.label5.grid_remove()
		self.entry4.grid_remove()
		self.cbutton1.grid_remove()
		self.cbutton2.grid_remove()
		self.button3.grid_remove()
		self.label6.grid_remove()
		self.fill9.grid_remove()
		self.fill10.grid_remove()
		self.fill11.grid_remove()
		self.label13.grid_remove()
		self.fill12.grid_remove()
		self.label14.grid_remove()
		self.entry9.grid_remove()
		self.searchbutton3.grid_remove()
		self.button5.grid_remove()
		self.label15.grid_remove()
		self.fill13.grid_remove()
		self.fill14.grid_remove()
		self.label16.grid_remove()
		self.entry10.grid_remove()
		self.searchbutton4.grid_remove()
		self.button6.grid_remove()
		self.label17.grid_remove()
		self.label18.grid_remove()
		self.fill15.grid_remove()
		self.fill16.grid_remove()
		self.fill17.grid_remove()
		self.fill18.grid_remove()

	def frame2(self):
		self.label7.grid()
		self.entry7.grid()
		self.searchbutton2.grid()
		self.label8.grid()
		self.button2.grid()
		self.label11.grid()
		self.fill5.grid()
		self.fill6.grid()
		self.fill7.grid()
		self.fill8.grid()
		self.msg1.set('')
		self.specpath.set('')

		self.labelfill1.grid_remove()
		self.labelfill2.grid_remove()
		self.labelfill3.grid_remove()
		self.labelfill4.grid_remove()
		self.labelfill5.grid_remove()
		self.labelfill6.grid_remove()
		self.labelfill7.grid_remove()
		self.labelfill8.grid_remove()
		self.label1.grid_remove()
		self.entry1.grid_remove()
		self.searchbutton1.grid_remove()
		self.button1.grid_remove()
		self.label10.grid_remove()
		self.label3.grid_remove()
		self.fill1.grid_remove()
		self.fill2.grid_remove()
		self.fill3.grid_remove()
		self.fill4.grid_remove()
		self.label4.grid_remove()
		self.label12.grid_remove()
		self.entry3.grid_remove()
		self.label5.grid_remove()
		self.entry4.grid_remove()
		self.cbutton1.grid_remove()
		self.cbutton2.grid_remove()
		self.button3.grid_remove()
		self.label6.grid_remove()
		self.fill9.grid_remove()
		self.fill10.grid_remove()
		self.fill11.grid_remove()
		self.label13.grid_remove()
		self.fill12.grid_remove()
		self.label14.grid_remove()
		self.entry9.grid_remove()
		self.searchbutton3.grid_remove()
		self.button5.grid_remove()
		self.label15.grid_remove()
		self.fill13.grid_remove()
		self.fill14.grid_remove()
		self.label16.grid_remove()
		self.entry10.grid_remove()
		self.searchbutton4.grid_remove()
		self.button6.grid_remove()
		self.label17.grid_remove()
		self.label18.grid_remove()
		self.fill15.grid_remove()
		self.fill16.grid_remove()
		self.fill17.grid_remove()
		self.fill18.grid_remove()

	def frame3(self):
		self.label4.grid()
		self.entry3.grid()
		self.label5.grid()
		self.entry4.grid()
		self.cbutton1.grid()
		self.cbutton2.grid()
		self.button3.grid()
		self.label12.grid()
		self.label6.grid()
		self.fill9.grid()
		self.fill10.grid()
		self.testsetpath.set('')
		self.testsetname.set('')
		self.spartamsg.set('')

		self.labelfill1.grid_remove()
		self.labelfill2.grid_remove()
		self.labelfill3.grid_remove()
		self.labelfill4.grid_remove()
		self.labelfill5.grid_remove()
		self.labelfill6.grid_remove()
		self.labelfill7.grid_remove()
		self.labelfill8.grid_remove()
		self.label1.grid_remove()
		self.entry1.grid_remove()
		self.searchbutton1.grid_remove()
		self.button1.grid_remove()
		self.label10.grid_remove()
		self.label3.grid_remove()
		self.fill1.grid_remove()
		self.fill2.grid_remove()
		self.fill3.grid_remove()
		self.fill4.grid_remove()
		self.label7.grid_remove()
		self.entry7.grid_remove()
		self.searchbutton2.grid_remove()
		self.label8.grid_remove()
		self.button2.grid_remove()
		self.label11.grid_remove()
		self.fill5.grid_remove()
		self.fill6.grid_remove()
		self.fill7.grid_remove()
		self.fill8.grid_remove()
		self.fill11.grid_remove()
		self.label13.grid_remove()
		self.fill12.grid_remove()
		self.label14.grid_remove()
		self.entry9.grid_remove()
		self.searchbutton3.grid_remove()
		self.button5.grid_remove()
		self.label15.grid_remove()
		self.fill13.grid_remove()
		self.fill14.grid_remove()
		self.label16.grid_remove()
		self.entry10.grid_remove()
		self.searchbutton4.grid_remove()
		self.button6.grid_remove()
		self.label17.grid_remove()
		self.label18.grid_remove()
		self.fill15.grid_remove()
		self.fill16.grid_remove()
		self.fill17.grid_remove()
		self.fill18.grid_remove()

	def frame4(self):
		self.fill11.grid()
		self.label13.grid()
		self.fill12.grid()
		self.label14.grid()
		self.entry9.grid()
		self.searchbutton3.grid()
		self.button5.grid()
		self.label15.grid()
		self.fill13.grid()
		self.fill14.grid()
		self.msg2.set('')
		self.ddlpath.set('')
		
		self.labelfill1.grid_remove()
		self.labelfill2.grid_remove()
		self.labelfill3.grid_remove()
		self.labelfill4.grid_remove()
		self.labelfill5.grid_remove()
		self.labelfill6.grid_remove()
		self.labelfill7.grid_remove()
		self.labelfill8.grid_remove()
		self.label1.grid_remove()
		self.entry1.grid_remove()
		self.searchbutton1.grid_remove()
		self.button1.grid_remove()
		self.label3.grid_remove()
		self.label10.grid_remove()
		self.fill1.grid_remove()
		self.fill2.grid_remove()
		self.fill3.grid_remove()
		self.fill4.grid_remove()
		self.label7.grid_remove()
		self.entry7.grid_remove()
		self.searchbutton2.grid_remove()
		self.label8.grid_remove()
		self.button2.grid_remove()
		self.label11.grid_remove()
		self.fill5.grid_remove()
		self.fill6.grid_remove()
		self.fill7.grid_remove()
		self.fill8.grid_remove()
		self.label4.grid_remove()
		self.entry3.grid_remove()
		self.label5.grid_remove()
		self.entry4.grid_remove()
		self.cbutton1.grid_remove()
		self.cbutton2.grid_remove()
		self.button3.grid_remove()
		self.label12.grid_remove()		
		self.label6.grid_remove()
		self.progbar.grid_remove()
		self.fill9.grid_remove()
		self.fill10.grid_remove()
		self.label16.grid_remove()
		self.entry10.grid_remove()
		self.searchbutton4.grid_remove()
		self.button6.grid_remove()
		self.label17.grid_remove()
		self.label18.grid_remove()
		self.fill15.grid_remove()
		self.fill16.grid_remove()
		self.fill17.grid_remove()
		self.fill18.grid_remove()

	def frame5(self):
		self.label16.grid()
		self.entry10.grid()
		self.searchbutton4.grid()
		self.button6.grid()
		self.label17.grid()
		self.label18.grid()
		self.fill15.grid()
		self.fill16.grid()
		self.fill17.grid()
		self.fill18.grid()
		self.msg3.set('')
		self.qexesrcpath.set('')

		self.labelfill1.grid_remove()
		self.labelfill2.grid_remove()
		self.labelfill3.grid_remove()
		self.labelfill4.grid_remove()
		self.labelfill5.grid_remove()
		self.labelfill6.grid_remove()
		self.labelfill7.grid_remove()
		self.labelfill8.grid_remove()
		self.label1.grid_remove()
		self.entry1.grid_remove()
		self.searchbutton1.grid_remove()
		self.button1.grid_remove()
		self.label3.grid_remove()
		self.label10.grid_remove()
		self.fill1.grid_remove()
		self.fill2.grid_remove()
		self.fill3.grid_remove()
		self.fill4.grid_remove()
		self.label7.grid_remove()
		self.entry7.grid_remove()
		self.searchbutton2.grid_remove()
		self.label8.grid_remove()
		self.button2.grid_remove()
		self.fill5.grid_remove()
		self.fill6.grid_remove()
		self.fill7.grid_remove()
		self.fill8.grid_remove()
		self.label4.grid_remove()
		self.label11.grid_remove()
		self.label12.grid_remove()
		self.entry3.grid_remove()
		self.label5.grid_remove()
		self.entry4.grid_remove()
		self.cbutton1.grid_remove()
		self.cbutton2.grid_remove()
		self.button3.grid_remove()
		self.label6.grid_remove()
		self.fill9.grid_remove()
		self.fill10.grid_remove()
		self.fill11.grid_remove()
		self.label13.grid_remove()
		self.fill12.grid_remove()
		self.label14.grid_remove()
		self.entry9.grid_remove()
		self.searchbutton3.grid_remove()
		self.button5.grid_remove()
		self.label15.grid_remove()
		self.fill13.grid_remove()
		self.fill14.grid_remove()

	def start_thread(self):
		self.button3['state'] = 'disable'
		self.entry3['state'] = 'disable'
		self.entry4['state'] = 'disable'
		self.cbutton1['state'] = 'disable'
		self.cbutton2['state'] = 'disable'
		self.b1['state'] = 'disable'
		self.b2['state'] = 'disable'
		self.b3['state'] = 'disable'
		self.b4['state'] = 'disable'
		self.b5['state'] = 'disable'		
		self.progbar.grid()
		self.progbar.start()
		print 'Initiating test execution process'
		self.label6.config(foreground='black')
		self.spartamsg.set('Connecting to HP ALM')
		self.q = Queue()
		self.secondary_process =Process(target=call_SpartaSQLServer,args=(self.continuerunflag.get(),self.defectflag.get(),self.q,self.testsetpath.get(),self.testsetname.get(),self.qcuser,self.qcpass,self.SQLserver))
		self.secondary_process.start()
		self.master.after(50, self.check_thread)

	def check_thread(self):
		if self.secondary_process.is_alive():
			if not self.q.empty():
				fromSparta = self.q.get()
				tcname='Executing Test Case '+fromSparta
				self.spartamsg.set(tcname)
			self.master.after(50, self.check_thread)
		else:
			if not self.q.empty():
				fromSparta = self.q.get()
				if '***ERROR***' in fromSparta:
					self.label6.config(foreground='red')
					self.spartamsg.set(fromSparta)
			else:
				self.label6.config(foreground='limegreen')
				self.spartamsg.set('Test Execution for Test Set "{0}" complete'.format(self.testsetname.get()))
			self.progbar.stop()
			self.progbar.grid_remove()
			self.button3['state'] = 'normal'
			self.entry3['state'] = 'normal'
			self.entry4['state'] = 'normal'
			self.cbutton1['state'] = 'normal'
			self.cbutton2['state'] = 'normal'
			self.b1['state'] = 'normal'
			self.b2['state'] = 'normal'
			self.b3['state'] = 'normal'
			self.b4['state'] = 'normal'
			self.b5['state'] = 'normal'

	def start_qexethread(self):
		self.button6['state'] = 'disable'
		self.entry10['state'] = 'disable'
		self.b1['state'] = 'disable'
		self.b2['state'] = 'disable'
		self.b3['state'] = 'disable'
		self.b4['state'] = 'disable'
		self.b5['state'] = 'disable'		
		# self.progbar.grid()
		# self.progbar.start()
		print 'Initiating test execution process'
		self.label18.config(foreground='black')
		self.msg3.set('Initiating test execution process')
		self.qexeq = Queue()
		self.secondary_qexeprocess =Process(target=call_Qexe,args=(self.qexesrcpath.get(),self.qexeq))
		self.secondary_qexeprocess.start()
		self.master.after(50, self.check_qexethread)

	def check_qexethread(self):
		if self.secondary_qexeprocess.is_alive():
			# print 'secondary process alive'
			if not self.qexeq.empty():
				qexetcname=self.qexeq.get()
				#print tcname
				self.msg3.set(qexetcname)
			self.master.after(50, self.check_qexethread)
		else:
			#print 'Sparta complete'
			if not self.qexeq.empty():
				self.qexetcname=self.qexeq.get()
			if '|||Error|||' in self.qexetcname:
				self.label18.config(foreground='red')
				self.msg3.set(self.qexetcname)
			else:
				self.label18.config(foreground='limegreen')
				self.msg3.set('Process complete')
			# self.progbar.stop()
			# self.progbar.grid_remove()
			self.button6['state'] = 'normal'
			self.entry10['state'] = 'normal'
			self.b1['state'] = 'normal'
			self.b2['state'] = 'normal'
			self.b3['state'] = 'normal'
			self.b4['state'] = 'normal'
			self.b5['state'] = 'normal'

	def getsrc(self):
		self.name= askopenfilename()
		self.srcpath.set(self.name)
		self.master.lift()

	def getsrc2(self):
		self.name= askopenfilename()
		self.specpath.set(self.name)
		self.master.lift()

	def getsrc3(self):
		self.name= askopenfilename()
		self.ddlpath.set(self.name)
		self.master.lift()

	def getsrc4(self):
		self.name= askopenfilename()
		self.qexesrcpath.set(self.name)
		self.master.lift()
	    
	def call_SPEQ(self):
		spec2q(self)
		self.message1='ETL Query generated : Filename - Output.txt  '
		self.msg1.set(self.message1)

	def call_PyRite(self):
		src= self.srcpath.get()
		tgt=self.despath.get()
		tgtname = pyrite(src,tgt)
		self.message='Test Cases Generated : Filename - TC_'+tgtname+'_TEST_CASES.xlsx'
		self.msg.set(self.message)

	def call_DDL(self):
		ddl(self.ddlpath.get(),self.dbuser,self.dbpass,self.dbhost,self.dbport,self.dbserv)
		self.message2='DDL Validation Completed '
		self.msg2.set(self.message2)


	def closePage(self):
		self.master.destroy()



setfailflag = False
failtc = False		

def crossDb_diff(x):
	global setfailflag
	global failtc
	if str(x[0]) == str(x[1]):
		return x[0]
	else:
		if not setfailflag:
			setfailflag = True
			failtc = True
		return '{} ---> {}'.format(*x)


def call_Qexe(qexefilepath,queue):
	global setfailflag
	global failtc
	print 'filepath is ',qexefilepath
	if os.path.exists('LOGS') == False:
		os.mkdir('LOGS')
	cwd = os.getcwd()
	log = open(r'SpartaRunTimeLog.txt','w')
	log.write('-----------------------------------------\n')
	log.write('Start time: '+str(datetime.now())+'\n')
	log.write('-----------------------------------------\n')
	# file = "Query_Input.xlsx"
	file = qexefilepath
	dfCred = pd.read_excel(file, sheet_name = 'Credentials', header = None)

	db = dfCred.iat[0,1]
	print 'DB Connection'
	log.write('DB Connection\n')
	if db == 'MYSQL':
		print 'MYSQL'
		log.write('MYSQL\n')
		dbusername =  dfCred.iat[10,0] #"root" 
		dbpassword = dfCred.iat[10,1] #"Password@2018"
		dbserver = dfCred.iat[10,2] #"localhost"
		dbschema = dfCred.iat[10,3] #"test"
		try:
			connection = MySQLdb.connect(dbserver,dbusername,dbpassword,dbschema)
			print 'Connected to MYSQL'
			log.write('Connected to MYSQL\n')
		except Exception as e:
			print '***ERROR ----- Connection to MYSQL Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'
			queue.put('|||Error||| Refer terminal for details')
			log.write('***ERROR ----- Connection to MYSQL Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
			log.close()
			return

	elif db == 'SQL Server':
		print 'SQL Server'
		log.write('SQL Server\n')
		driver = dfCred.iat[4,0]
		server = dfCred.iat[4,1]
		database = dfCred.iat[4,2]
		userId = dfCred.iat[4,3]
		password = dfCred.iat[4,4]
		port = dfCred.iat[4,5]
		try:
			connection = pyodbc.connect(driver=driver,server=server,database=database,trusted_connection='yes') #Not sure if object returned will be compatible with pandas read_sql function
			print 'Connected to to SQL Server'
			log.write('Connected to SQL Server\n')
		except  Exception as e1:
			print '***ERROR ----- Connection to SQL Server Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'
			queue.put('|||Error||| Refer terminal for details')
			log.write('***ERROR ----- Connection to SQL Server Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
			log.close()
			return

	elif db == 'Oracle':
		print 'Oracle'
		log.write('Oracle\n')
		dbconnectsting = str(dfCred.iat[7,0]+'/'+dfCred.iat[7,1]+'@'+dfCred.iat[7,2]+':'+dfCred.iat[7,3]+'/'+dfCred.iat[7,4])
		try:
			connection = cx_Oracle.connect(dbconnectsting)
			print 'Connected to Oracle'
			log.write('Connected to Oracle\n')
		except  cx_Oracle.DatabaseError as e1:
			print '***ERROR ----- Connection to Oracle Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'
			queue.put('|||Error||| Refer terminal for details')
			log.write('***ERROR ----- Connection to Oracle Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
			log.close()
			return

	else:
		print 'Invalid DB Type. Check Credentials'
		queue.put('|||Error||| Refer terminal for details')
		log.write('***ERROR ----- Invalid DB Type. Check Credentials ----- ERROR ***\n')
		log.close()
		return

	df = pd.read_excel(file, sheet_name = 'Query_Input')

	passcount = 0
	failcount = 0
	notcompletedcount = 0
	TestCaseTypeList = df.loc[:,'Type of Test Case'].tolist()
	crossDB = 'Cross DB Validation - Count' in TestCaseTypeList or 'Cross DB Validation - Data' in TestCaseTypeList
	if crossDB:
		print 'Cross DB Connection'
		log.write('Cross DB Connection\n')

		dfCrossDb = pd.read_excel(file, sheet_name = 'Cross DB',header = None)
		crossDbSrc = dfCrossDb.iat[1,1]
		crossDbTgt = dfCrossDb.iat[17,1]
		if crossDbSrc == 'MYSQL':
			print 'Src : MYSQL'
			log.write('Src : MYSQL\n')
			dbusername =  dfCrossDb.iat[11,0] #"root" 
			dbpassword = dfCrossDb.iat[11,1] #"Password@2018"
			dbserver = dfCrossDb.iat[11,2] #"localhost"
			dbschema = dfCrossDb.iat[11,3] #"test"
			try:
				srcConnection = MySQLdb.connect(dbserver,dbusername,dbpassword,dbschema)
				print 'Connected to MYSQL'
				log.write('Connected to MYSQL\n')
			except Exception as e:
				print '***ERROR ----- Connection to MYSQL Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to MYSQL Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		elif crossDbSrc == 'SQL Server':
			print 'Src : SQL Server'
			log.write('Src : SQL Server\n')
			driver = dfCrossDb.iat[5,0]
			server = dfCrossDb.iat[5,1]
			database = dfCrossDb.iat[5,2]
			userId = dfCrossDb.iat[5,3]
			password = dfCrossDb.iat[5,4]
			port = dfCrossDb.iat[5,5]
			try:
				srcConnection = pyodbc.connect(driver=driver,server=server,database=database,trusted_connection='yes') #Not sure if object returned will be compatible with pandas read_sql function
				print 'Connected to SQL Server'
				log.write('Connected to SQL Server\n')
			except  Exception as e1:
				print '***ERROR ----- Connection to SQL Server Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'				
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to SQL Server Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		elif crossDbSrc == 'Oracle':
			print 'Src : Oracle'
			log.write('Src : Oracle\n')
			dbconnectsting = str(dfCrossDb.iat[8,0]+'/'+dfCrossDb.iat[8,1]+'@'+dfCrossDb.iat[8,2]+':'+dfCrossDb.iat[8,3]+'/'+dfCrossDb.iat[8,4])
			try:
				srcConnection = cx_Oracle.connect(dbconnectsting)
				print 'Connected to Oracle'
				log.write('Connected to Oracle\n')
			except  cx_Oracle.DatabaseError as e1:
				print '***ERROR ----- Connection to Oracle Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***'				
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to Oracle Failed CHECK YOUR DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		else:
			print 'Invalid DB Type. Check Cross DB Credentials'
			log.write('Invalid DB Type. Check Cross DB Credentials')
			return

		if crossDbTgt == 'MYSQL':
			print 'Tgt : MYSQL'
			log.write('Tgt : MYSQL\n')
			dbusername =  dfCrossDb.iat[27,0] #"root" 
			dbpassword = dfCrossDb.iat[27,1] #"Password@2018"
			dbserver = dfCrossDb.iat[27,2] #"localhost"
			dbschema = dfCrossDb.iat[27,3] #"test"
			try:
				tgtConnection = MySQLdb.connect(dbserver,dbusername,dbpassword,dbschema)
				print 'Connected to MYSQL'
				log.write('Connected to MYSQL\n')
			except Exception as e:
				print '***ERROR ----- Connection to MYSQL Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***'
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to MYSQL Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		elif crossDbTgt == 'SQL Server':
			print 'Tgt : SQL Server'
			log.write('Tgt : SQL Server\n')
			driver = dfCrossDb.iat[21,0]
			server = dfCrossDb.iat[21,1]
			database = dfCrossDb.iat[21,2]
			userId = dfCrossDb.iat[21,3]
			password = dfCrossDb.iat[21,4]
			port = dfCrossDb.iat[21,5]
			try:
				tgtConnection = pyodbc.connect(driver=driver,server=server,database=database,trusted_connection='yes') #Not sure if object returned will be compatible with pandas read_sql function
				print 'Connected to SQL Server'
				log.write('Connected to SQL Server\n')
			except  Exception as e1:
				print '***ERROR ----- Connection to SQL Server Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***'				
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to SQL Server Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		elif crossDbTgt == 'Oracle':
			print 'Tgt : Oracle\n'
			dbconnectsting = str(dfCrossDb.iat[24,0]+'/'+dfCrossDb.iat[24,1]+'@'+dfCrossDb.iat[24,2]+':'+dfCrossDb.iat[24,3]+'/'+dfCrossDb.iat[24,4])
			try:
				tgtConnection = cx_Oracle.connect(dbconnectsting)
				print 'Connected to Oracle'
				log.write('Connected to Oracle\n')
			except  Exception as e1:
				print '***ERROR ----- Connection to Oracle Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***'				
				queue.put('|||Error||| Refer terminal for details')
				log.write('***ERROR ----- Connection to Oracle Failed CHECK YOUR CROSS DB CREDENTIALS ----- ERROR ***\n')
				log.close()
				return

		else:
			print 'Invalid DB Type. Check Cross DB Credentials'
			queue.put('|||Error||| Refer terminal for details')
			log.write('Invalid DB Type. Check Cross DB Credentials\n')
			log.close()
			return


	for i,row in df.iterrows():
		log = open(r'\\CTSINTBMNTAPPA\Innovation\Sriram\SPARTA\SpartaRunTimeLog.txt','a')
		table_name = (df['Table Name'][i])
		query = str(df['Source Query'][i])
		queryType = str(df['Type of Test Case'][i])
		queue.put('\nExecuting TC : '+str(df['Test Case Name'][i]))
		querystring = sqlparse.format(query, reindent = True, keyword_case = 'upper')
		print 'Executing TC: ', str(df['Test Case Name'][i])
		log.write ('Executing TC: '+ str(df['Test Case Name'][i])+'\n')
		crossDbData = False
		status = ""
		setfailflag = False
		failtc = False
		crossDB = 'Cross DB Validation - Count' in queryType or 'Cross DB Validation - Data' in queryType
		if not crossDB:
			try:
				result = pd.read_sql(sql = querystring,con=connection)
			except:
				print 'Query Error. Please check query...!!!!'
				log.write('Query Error. Please check query...!!!!\n')
				log.write('Status-->Not Completed\n')
				log.close()
				df.loc[i,'Syntax Error']="Error"
				status = 'Not Completed'
				notcompletedcount = notcompletedcount + 1
				df.loc[i,'Execution Status'] = status
				continue
			if 'Data Validation' in queryType:
				if result.empty:
					df.loc[i,'Records Returned'] = 0
					status = "Passed"
					passcount = passcount + 1
					log.write('Status-->Passed\n')
					log.close()
				else:
					df.loc[i,'Records Returned'] = result.shape[0]
					status = "Failed"
					failcount = failcount + 1
					log.write('Status-->Failed\n')
					log.close()
			elif 'Count Validation' in queryType:
				if result.shape == (2,2):
					counts = result.iloc[:,1:2]
					srccnt = counts.iat[0,0]
					tgtcnt = counts.iat[1,0]
					df.loc[i,'Records Returned'] = 'SRC - '+str(srccnt)+', TGT - '+str(tgtcnt)
					counts = result.iloc[:,1:2]
					if srccnt == tgtcnt:
						status = 'Passed'
						passcount = passcount + 1
						log.write('Status-->Passed\n')
						log.close()
					else:
						status = 'Failed'
						failcount = failcount + 1
						log.write('Status-->Failed\n')
						log.close()
				elif result.empty:	# count minus query returns empty set
					df.loc[i,'Records Returned'] = 0
					status = "Passed"
					passcount = passcount + 1
					log.write('Status-->Passed\n')
					log.close()
				else:
					df.loc[i,'Records Returned'] = result.shape[0]
					status = "Failed"
					failcount = failcount + 1
					log.write('Status-->Failed\n')
					log.close()
		elif crossDB:
			tgtQuery = str(df['Target Query'][i])
			tgtQueryString = sqlparse.format(tgtQuery, reindent = True, keyword_case = 'upper')
			try:
				srcResult = pd.read_sql(sql = querystring,con=srcConnection)
			except:
				print 'Source Query Error. Please check query...!!!!'
				log.write('Source Query Error. Please check query...!!!!\n')
				log.write('Status-->Not Completed\n')
				log.close()
				df.loc[i,'Syntax Error']="Error"
				status = 'Not Completed'
				notcompletedcount = notcompletedcount + 1
				df.loc[i,'Execution Status'] = status
				continue
			try:
				tgtResult = pd.read_sql(sql = tgtQueryString,con=tgtConnection)
			except:
				print 'Target Query Error. Please check query...!!!!'
				log.write('Target Query Error. Please check query...!!!!\n')
				log.write('Status-->Not Completed\n')
				log.close()
				df.loc[i,'Syntax Error']="Error"
				status = 'Not Completed'
				notcompletedcount = notcompletedcount + 1
				df.loc[i,'Execution Status'] = status
				continue
			if 'Count' in queryType:	#Cross DB Count
				srccnt = srcResult.iat[0,0]
				tgtcnt = tgtResult.iat[0,0]
				df.loc[i,'Records Returned'] = 'SRC - '+str(srccnt)+', TGT - '+str(tgtcnt)
				if srccnt == tgtcnt:
					status = 'Passed'
					passcount = passcount + 1
					log.write('Status-->Passed\n')
					log.close()
				else:
					status = 'Failed'
					failcount = failcount + 1
					log.write('Status-->Failed\n')
					log.close()
			elif 'Data' in queryType:	#Cross DB data
				crossDbData = True
				diff_panel = pd.Panel(dict(df1=srcResult,df2=tgtResult))
				diff_output = diff_panel.apply(crossDb_diff, axis=0)
				df.loc[i,'Records Returned'] = 'SRC - '+str(srcResult.shape[0])+', TGT - '+str(tgtResult.shape[0])
				if not failtc:
					status = 'Passed'
					passcount = passcount + 1
					log.write('Status-->Passed\n')
					log.close()
				else:
					status = 'Failed'
					failcount = failcount + 1
					log.write('Status-->Failed\n')
					log.close()

		df.loc[i,'Execution Status'] = status
		outputname = cwd+'\LOGS\log_'+str(table_name)+'_'+str(df['Test Case Name'][i])+'_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx'
		writer = pd.ExcelWriter(outputname)
		if crossDB:
			if crossDbData:
				srcResult.to_excel(writer,'SOURCE RESULT')
				tgtResult.to_excel(writer,'TARGET RESULT')
				diff_output.to_excel(writer,'DATA COMPARISON')
			else:
				srcResult.to_excel(writer,'SOURCE OUTPUT',index=False)
				tgtResult.to_excel(writer,'TARGET OUTPUT',index=False)
		else:
			result.to_excel(writer,'OUTPUT',index=False)
		writer.save()
	writer1 = pd.ExcelWriter(cwd+'\LOGS\Status_Sheet'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')
	df.to_excel(writer1,'Status',index = False)
	workbook = writer1.book
	worksheet= writer1.sheets['Status']
	
	worksheet.set_column('A:A',6)
	worksheet.set_column('B:B',20)
	worksheet.set_column('C:C',33)
	worksheet.set_column('D:D',23)
	worksheet.set_column('E:E',180)
	worksheet.set_column('F:F',18)
	worksheet.set_column('G:G',13)
	worksheet.set_column('H:H',12)
	worksheet.set_column('I:I',10)
	worksheet.set_column('J:J',8)
	worksheet.set_column('K:K',15)
	worksheet.set_column('L:L',15)
	
	writer1.save()
	log = open(r'\\CTSINTBMNTAPPA\Innovation\Sriram\SPARTA\SpartaRunTimeLog.txt','a')
	log.write('-----------------------------------------\n')
	log.write('End time: '+str(datetime.now())+'\n')
	log.write('-----------------------------------------\n')
	log.close()
	connection.close()
	if crossDB:
		srcConnection.close()
		tgtConnection.close()

	fig = plt.figure(figsize=(10,7))
	statusgraph = fig.add_subplot(1,1,1)

	statusgraph.tick_params(axis='x',which='both',bottom='off',top='off',labelbottom='off')

	index = 1
	bar_width = 0.25
	opacity = 0.4

	rects1 = statusgraph.bar(index, passcount, bar_width,alpha=opacity,color='g',label='Passed')
	rects2 = statusgraph.bar(index+bar_width+bar_width, failcount, bar_width,alpha=opacity,color='r',label='Failed')
	rects3 = statusgraph.bar(index+bar_width+bar_width+bar_width+bar_width, notcompletedcount, bar_width,alpha=opacity,color='b',label='Not Completed')

	statusgraph.set_xlabel('Status')
	statusgraph.set_ylabel('Test Case Count')
	statusgraph.legend()
	statusgraph.set_title('Count Plot')
	fig.savefig(cwd+'\LOGS\Dashboard'+time.strftime("%Y%m%d-%H%M%S")+'.png',bbox_inches='tight')   # save the figure to file
	plt.close(fig) 
	print 'Execution Completed'


def call_SpartaSQLServer(continuerunflag,defectflag,queue,testsetpath,testsetname,qcuser,qcpass,SQLServer):
	
	path_to_testset=testsetpath # "Subject\DOL_DAY_1_REGRESSION"
	testset_name =testsetname # 'DEMO_REGRESSION'
	hp_alm_url = 'http://staap1509.r02.xlgs.local:8080/qcbin'
	hp_alm_user = qcuser
	hp_alm_pass = qcpass
	hp_alm_project_name = 'Data_Conforming'
	hp_alm_domain_name = 'XLI'
	try:
		hp_alm_connection_obj = win32com.client.Dispatch("TDApiOle80.TDConnection") 
		hp_alm_connection_obj.InitConnection(hp_alm_url) 
		hp_alm_connection_obj.Login(hp_alm_user,hp_alm_pass)
		hp_alm_connection_obj.Connect(hp_alm_domain_name,hp_alm_project_name)
	except Exception as e:
		print '!!!! ALM Connection Object Creation Failed -->\n',e
		queue.put('***ERROR*** Refer Terminal for details')
		return

	if hp_alm_connection_obj.Connected is True:
		print 'Connection to HP ALM Successful'
	else:
		print "************ Connection to QC failed. Check your credentials! ***********"
		queue.put('***ERROR*** Refer Terminal for details')
		return
	print("Path to testset: "+path_to_testset+"\nTestset_name: "+testset_name)
	tsFolder = hp_alm_connection_obj.TestSetTreeManager.NodeByPath(path_to_testset)
	tsList = tsFolder.FindTestSets(testset_name)
	if (len(tsList)) == 0:
		print("No Tests found")
	for ts in tsList:
		print 'TEST SET NAME IS ',ts.Name
		if ts.Name ==testset_name:
			test_set_object = ts
	print test_set_object.Name
	if test_set_object.Name==testset_name:
		print 'SPARTA IS EXECUTING....',test_set_object.Name,testset_name
	
		TSTestFact = test_set_object.TSTestFactory
		tsFilter = TSTestFact.Filter
		tsFilter["TC_CYCLE_ID"] = test_set_object.ID
		# tsFilter["CY_CYCLE"] = test_set_object.Name
		# print test_set_object.Name
		testList = TSTestFact.NewList(tsFilter.Text)		

		passcount=0
		failcount=0
		noruncount=0
		notcompletedcount=0
		tc_time=[]
		tc_id = []
		testcaseList = []

		if continuerunflag:
			for test_case in testList :
				if test_case.Status == 'No Run' or test_case.Status == 'Not Completed':
					testcaseList.append(test_case)
			print '# of TC in continue mode - ',len(testcaseList),'\n# of total TC - ',len(testList)
			testList = testcaseList

		for test_case in testList :
			s=time.time()
			testername= test_case.Field("TC_ACTUAL_TESTER")
			relname = test_set_object.Field("CY_USER_02")
			applname= test_case.Field("TS_USER_03")
			testcasepriority= test_case.Field("TS_USER_01")
			start_time_test = time.time();
			queue.put(test_case.Name)
			print test_case.ID,test_case.Name

			fail_flag = False
			queryflag = False
			syntaxflag = False
			connflag = False
			cantdocount= False

			newItem = test_case.RunFactory.AddItem(None)   # newItem == Run Object 
			newItem.Status = 'No Run' 
			newItem.Name = 'Run '+ datetime.datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
			newItem.SetField("RN_USER_06","Automated")
			newItem.Post() 
			try:
				newItem.CopyDesignSteps() 
			except:
				print 'Unknown Error'
				continue
			newItem.Post() 
			steps = newItem.StepFactory.NewList("")
			
			for step in steps :
				result_string =""
				step = steps[0]
				step.Status = "Not Completed" 
				soup = BeautifulSoup(step.Field('ST_DESCRIPTION'),"html.parser")
				for script in soup(["script", "style"]):
				    script.extract()    # rip it out
				text = soup.get_text(strip=True,separator='\n')
				text = text.replace("Query:","").replace(u'\xa0',' ')
				text=text.replace(u'\xc2\xa0', ' ')
				text=text.replace(u'\u00A0', ' ')
				text=text.replace(u'\r\n', '')
				text = text.replace(";","")
				text = text.strip('"')
				text=sqlparse.format(text, reindent=True,keyword_case='upper')
				if not text == "" :
					try:
						con = pyodbc.connect(driver='{SQL Server}',server=server,trusted_connection='yes')
					except Exception as e1:
						connflag = True
						print e1
						print '***ERROR ----- Connection to Oracle Failed CHECK YOUR DB CREDENTIALS ----------ERROR ************'
						break
					cur = con.cursor()

					soup = BeautifulSoup(test_case.Name,"html.parser")
					tcname = soup.get_text(strip=True,separator='\n')
					printable = set(string.printable)
					tcname = filter(lambda x: x in printable, tcname)
					print tcname
					tcname= tcname.strip().replace(" ", "")
					print tcname
					tcname = tcname.upper()

					if '_DATACOMP' in tcname:
						try:
							datacompdf = pd.read_sql(sql=text, con=con)
						except Exception as e:
							print "!!!ERROR!!!\n",e
							syntaxflag = True
							continue
						datacompdf = datacompdf.iloc[:,2::3]
						checklist = datacompdf.columns.values.tolist()
						failedcol = list()
						for col in checklist:
							sample = list(datacompdf[col])
							if 'FAIL' in sample:
								failedcol.append(col)
						if not failedcol:  # empty failed col list
							step.Status = "Passed"
							step.SetField("ST_ACTUAL", "No failed columns")
							passcount = passcount +1
						else:
							step.Status = "Failed"
							step.SetField("ST_ACTUAL", "The following columns failed :"+str(failedcol))
							failcount = failcount +1
							fail_flag = True
					else:
						try:
							
							cur.execute("ALTER SESSION SET NLS_DATE_FORMAT = 'DD-MON-YYYY HH24:MI:SS' NLS_TIMESTAMP_FORMAT = 'DD-MON-YYYY HH.MI.SSXFF AM'")
							cur.execute(text)

						except cx_Oracle.DatabaseError as e:
							error, = e.args
							print 'Error Code',error.code

							print '***ORACLE SYNTAX ERROR :\n\t {0}\t {1}'.format(error.message,error.context)
							syntaxflag = True
							continue

						except MemoryError as memerr:
							syntaxflag = True
							print '***Memory Error'
							continue

						records_returned = cur.fetchall()
						if os.path.exists('H:/Desktop/LOGS')==False:
							os.mkdir('H:/Desktop/LOGS')
						try:
							pd.read_sql(sql=text, con=con) \
							.to_csv('H:/Desktop/LOGS/'+test_case.Name+'LOG.csv', index=False)
						except cx_Oracle.DatabaseError as e:
							error, = e.args
							print 'Error Code',error.code

							print '***ORACLE SYNTAX ERROR :\n\t {0}\t {1}'.format(error.message,error.context)
							syntaxflag = True
							continue

						except MemoryError as memerr:
							syntaxflag = True
							continue
						defect_ids = ""

						# should we comment the next line out, since wer updating the ST_ACTUAL in the if-else clause below
						# step.SetField("ST_ACTUAL", "Records Returned :"+str(cur.rowcount)) # Wohhooooo!!!!
						if cur.rowcount == 2:
							firstrow=records_returned[0]
							secondrow=records_returned[1]
							try:
								srccount = firstrow[1]
							except:
								syntaxflag = True
								continue

							tgtcount=secondrow[1]
							try:
								srccount-tgtcount
							except:
								cantdocount = True
								
						
						#for data test cases 
						if cur.rowcount == 0 and '_DATA' in tcname  :
							step.Status = "Passed"
							step.SetField("ST_ACTUAL", "Data match between SRC and TGT by minus query. Records Returned :"+str(cur.rowcount)) # Wohhooooo!!!!
							passcount = passcount +1
						# for count minus test cases
						elif cur.rowcount == 0 and '_COUNT' in tcname  :
							step.Status = "Passed"
							step.SetField("ST_ACTUAL", "Count match between SRC and TGT tables by minus query. Records Returned :"+str(cur.rowcount)) # Wohhooooo!!!!
							passcount = passcount +1

						#for count test cases
						elif (cantdocount == False and '_COUNT' in tcname and cur.rowcount == 3 ) :
							print 'Check Sum Query Encountered...!!!!'
							firstrow=records_returned[0]
							secondrow=records_returned[1]
							thirdrow=records_returned[2]
							try:
								srccount = firstrow[1]
							except:
								syntaxflag = True
								continue

							tgtcount=secondrow[1]
							checksum=thirdrow[1]
							print srccount,tgtcount,checksum
							try:
								
								diff = max(tgtcount,srccount) - (min(tgtcount,srccount)+checksum)

							except Exception, e:
								traceback.print_exc()
								syntaxflag = True
								continue

							if diff == 0:
								step.Status = "Passed"
								step.SetField("ST_ACTUAL", "Counts match between SRC and TGT tables by union all query. Matching Record Count: "+str(srccount)) # Wohhooooo!!!!
								passcount = passcount +1
							else:
								step.Status = "Failed"
								step.SetField("ST_ACTUAL", "Counts don't match between SRC and TGT tables. ") # Wohhooooo!!!!
								failcount = failcount +1
								fail_flag = True


						elif (cantdocount == False and cur.rowcount == 2 and srccount-tgtcount==0) :
							step.Status = "Passed"
							step.SetField("ST_ACTUAL", "Counts match between SRC and TGT tables by union all query. Matching Record Count: "+str(srccount)) # Wohhooooo!!!!
							passcount = passcount +1

						#for failed count test cases union all query
						elif (cantdocount == False and cur.rowcount == 2 and srccount-tgtcount!=0) :
							step.Status = "Failed"
							step.SetField("ST_ACTUAL", "Counts don't match between SRC and TGT tables")
							failcount = failcount +1
							fail_flag = True
						# count minus failure *******should come after count union failure*******
						elif cur.rowcount >= 1 and '_COUNT' in tcname  :
							step.Status = "Failed"
							step.SetField("ST_ACTUAL", "Counts don't match between SRC and TGT tables.")
							failcount = failcount +1
							fail_flag = True
						# for DDL/Duplicate/other queries returning 0 records
						elif cur.rowcount == 0 :
							step.Status = "Passed"
							step.SetField("ST_ACTUAL", "Data is matching. Records Returned :"+str(cur.rowcount)) # Wohhooooo!!!!
							passcount = passcount +1

						#for failing data test cases where rowcount >= 1
						elif (cantdocount == True and cur.rowcount >= 1) :
							step.Status = "Failed"
							step.SetField("ST_ACTUAL", "Data mismatch between SRC and TGT tables. Mismatching Record Count :"+str(cur.rowcount))
							failcount = failcount +1
							fail_flag = True

						else :
							step.Status = "Failed"
							step.SetField("ST_ACTUAL", "Data Mismatch. Records Returned :"+str(cur.rowcount))

							failcount = failcount +1
							fail_flag = True

					if defectflag and fail_flag == True:
						for result in records_returned:
							tup_str = ",".join([str(x) for x in result])
							result_string =  result_string + "\n"+tup_str;
							step.SetField("ST_ACTUAL", "Rows returned :"+str(cur.rowcount) ) # Wohhooooo!!!!
							# step.SetField("ST_ACTUAL", "Rows returned :"+str(cur.rowcount) + "Records :<br>" +result_string ) # Wohhooooo!!!!
						
						print applname,testername,relname
						bugfac = hp_alm_connection_obj.BugFactory;
						bug = bugfac.AddItem(None);
						# Found in Application  char
						bug.SetField("BG_USER_01", applname)

						# Assigned To   char
						bug.SetField("BG_RESPONSIBLE", testername)

						# Status    char.
						bug.SetField("BG_STATUS", "New")

						# Phase Detected    char
						bug.SetField("BG_USER_34", "Systems Test")

						# Release Project Name  char
						bug.SetField("BG_USER_30", relname)

						# Priority  char
						if testcasepriority=='Low':
							bug.SetField("BG_PRIORITY", "2-Medium")
						elif testcasepriority=='Medium':
							bug.SetField("BG_PRIORITY", "3-High")
						elif testcasepriority=='High':
							bug.SetField("BG_PRIORITY", "4-Very High")

						# Found By  char
						bug.SetField("BG_USER_06", "System Testing")

						# Severity  char
						if testcasepriority=='Low':
							bug.SetField("BG_SEVERITY", "4-Minor")
						elif testcasepriority=='Medium':
							bug.SetField("BG_SEVERITY", "3-Major")
						elif testcasepriority=='High':
							bug.SetField("BG_SEVERITY", "2-Critical")

						# dbconnectstingry   char
						bug.SetField("BG_SUMMARY", "Test Case : "+test_case.Name+" failed with records returend count :"+str(cur.rowcount))

						# Description   memo
						bug.SetField("BG_DESCRIPTION", 'DEFECT : Data mismatch found.'+" Please refer attachement for mismatching records")

						bug.post()

						attachfact=bug.Attachments;
						attachObj = attachfact.AddItem(None)
						attachObj.Description = "Failed Records"
						attachObj.Filename = "H:\\Desktop\\LOGS\\"+test_case.Name+"LOG.csv"
						attachObj.Type=1
						attachObj.Post()
						defect_ids = defect_ids + " "+str(bug.Field("BG_BUG_ID"))
					else:
						print 'Defect Creation disabled'
				cur.close()
				con.close()

				
				try:
					step.post() 
				except:
					print 'Trying to Re-connect to ALM'
					hp_alm_connection_obj.InitConnection(hp_alm_url) 
					hp_alm_connection_obj.Login(hp_alm_user,hp_alm_pass)
					hp_alm_connection_obj.Connect(hp_alm_project_name,hp_alm_domain_name) 
					print 'Re-connection to HP ALM Established...!!!!'
					step.post()


			if connflag:
				break
			if fail_flag :
				newItem.Status = "Failed"

			elif queryflag:
				newItem.Status = "Not Completed"
				steps = newItem.StepFactory.NewList("")
				for step in steps:
					step.Status = "Not Completed"
				step.SetField("ST_ACTUAL", "Long Running Query. Please optimise Query")
				step.post()

				notcompletedcount = notcompletedcount +1

			elif syntaxflag:
				newItem.Status = "Not Completed"
				steps = newItem.StepFactory.NewList("")
				for step in steps:
					step.Status = "Not Completed"
				step.SetField("ST_ACTUAL", "Syntax Error. Check Query")
				step.post()

				notcompletedcount = notcompletedcount +1
			else :
				newItem.Status = "Passed"
			newItem.Post()
			tc_time.append(time.time()-s)
			tc_id.append(test_case.Name)

		print tc_time

		fig = plt.figure(figsize=(12,7))
		statusgraph = fig.add_subplot(1,2,1)

		statusgraph.tick_params(axis='x',which='both',bottom='off',top='off',labelbottom='off')

		index = 1
		bar_width = 0.25
		opacity = 0.4

		rects1 = statusgraph.bar(index, passcount, bar_width,alpha=opacity,color='g',label='Passed')
		rects2 = statusgraph.bar(index+bar_width+bar_width, failcount, bar_width,alpha=opacity,color='r',label='Failed')
		rects3 = statusgraph.bar(index+bar_width+bar_width+bar_width+bar_width, notcompletedcount, bar_width,alpha=opacity,color='b',label='Not Completed')

		statusgraph.set_xlabel('Status')
		statusgraph.set_ylabel('Test Case Count')
		statusgraph.legend()
		statusgraph.set_title('Count Plot')

		timegraph = fig.add_subplot(1,2,2)

		labels = tc_id
		sizes = tc_time
		total = sum(tc_time)

		timegraph.pie(sizes, autopct = lambda(p): '{:.2f} secs'.format(p * total / 100), shadow=True, startangle=90)
		timegraph.set_title('Time Plot')
		timegraph.legend(labels)

		plt.tight_layout()
		plt.show()


def pyrite(src,tgt):
	highlight = NamedStyle(name="highlight")
	highlight.font = Font(bold=True, size=13, color="0000FF")
	bd = Side(style='thick', color="000000")
	highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
	highlight.alignment = Alignment(horizontal='center')

	srcfile=src	
	wb = openpyxl.load_workbook(srcfile)
	wbtgt = openpyxl.Workbook()
	wb.active
	sheet = wb.get_sheet_by_name('MAPPING')
	sheet2 = wb.get_sheet_by_name('PARAMETERS')
	sheettgt = wbtgt.active
	targetTable = sheet2['B4'].value
	desfile='TC_'+targetTable+'_TEST_CASES.xlsx'
	app = sheet2['B2'].value
	sub = sheet2['B3'].value
	tcDetails = {}

	wbtgt.add_named_style(highlight)

	for y in range(1,16):
		sheettgt.cell(row=1,column=y).style = highlight

	sheettgt.cell(row=1,column=1).value="Application"
	sheettgt.cell(row=1,column=2).value="Subject"
	sheettgt.cell(row=1,column=3).value="Description"
	sheettgt.cell(row=1,column=4).value="Test Name"
	sheettgt.cell(row=1,column=5).value="Test Case Status"
	sheettgt.cell(row=1,column=6).value="Step Name "
	sheettgt.cell(row=1,column=7).value="Step Description "
	sheettgt.cell(row=1,column=8).value="Expected Results"
	sheettgt.cell(row=1,column=9).value="Priority"
	sheettgt.cell(row=1,column=10).value="Regression Indicator"
	sheettgt.cell(row=1,column=11).value="Automation Indicator"
	sheettgt.cell(row=1,column=12).value="System Function"
	sheettgt.cell(row=1,column=13).value="Type"
	sheettgt.cell(row=1,column=14).value="Regression Category"
	sheettgt.cell(row=1,column=15).value="Regression Status"


	for row in range(1,sheet.max_row + 1):
		if((str(sheet['A' +str(row)].value)).isdigit()): 
			tc_count = int(sheet['A' +str(row)].value)
			tcDetails["sourceTable"] = sheet['B' +str(row)].value
			tcDetails["sourceColumn"] = sheet['C' +str(row)].value
			tcDetails["transformationLogic"] = sheet['E' +str(row)].value
			tcDetails["targetColumn"] = sheet['F' +str(row)].value
			tcDetails["regind"] = sheet['J' +str(row)].value
			tcDetails["regcat"] = sheet['K' + str(row)].value
			tcDetails["regstat"] = sheet['L' + str(row)].value
			
			for y in range(1,16):
				sheettgt.cell(row=row,column=y).border = Border(left=bd, right=bd)

			sheettgt.cell(row=row,column=1).value =app
			sheettgt.cell(row=row,column=2).value =sub
			sheettgt.cell(row=row,column=3).value ="Verification of "+tcDetails["targetColumn"]+" column in "+targetTable+" table"
			sheettgt.cell(row=row,column=4).value =targetTable+'_'+tcDetails["targetColumn"]+'_TC00'+str(tc_count)
			sheettgt.cell(row=row,column=5).value ="Complete"
			sheettgt.cell(row=row,column=6).value ="Step 1"
			sheettgt.cell(row=row,column=7).value ="Verify whether "+tcDetails["targetColumn"]+" column in "+targetTable+" table is populated as per the below logic : \n"+tcDetails["transformationLogic"]
			sheettgt.cell(row=row,column=8).value =tcDetails["targetColumn"]+" column in "+targetTable+" table is populated as per the mapping logic"
			sheettgt.cell(row=row,column=9).value ="Medium"
			sheettgt.cell(row=row,column=10).value =tcDetails["regind"]
			sheettgt.cell(row=row,column=11).value ="Automated"
			sheettgt.cell(row=row,column=12).value ="Workflow"
			sheettgt.cell(row=row,column=13).value ="Manual"
			sheettgt.cell(row=row,column=14).value =tcDetails["regcat"]
			sheettgt.cell(row=row,column=15).value =tcDetails["regstat"]

	tc_count=tc_count+1
	row=row+1
	sheettgt.cell(row=row,column=1).value =app
	sheettgt.cell(row=row,column=2).value =sub
	sheettgt.cell(row=row,column=3).value ="Verification of Successful run of Autosys job to load "+targetTable+" table"
	sheettgt.cell(row=row,column=4).value ="AUTOSYS_"+targetTable+'_TC00'+str(tc_count)
	sheettgt.cell(row=row,column=5).value ="Complete"
	sheettgt.cell(row=row,column=6).value ="Step 1"
	sheettgt.cell(row=row,column=7).value ="Verify whether the Autosys job was run successfully to load "+targetTable+" table"
	sheettgt.cell(row=row,column=8).value ="The Specified autosys job should run successfully to load "+targetTable+" table"
	sheettgt.cell(row=row,column=9).value ="Medium"
	sheettgt.cell(row=row,column=10).value ='No'
	sheettgt.cell(row=row,column=11).value ="Automated"
	sheettgt.cell(row=row,column=12).value ="Workflow"
	sheettgt.cell(row=row,column=13).value ="Manual"
	sheettgt.cell(row=row,column=14).value =tcDetails["regcat"]
	sheettgt.cell(row=row,column=15).value =tcDetails["regstat"]
	for y in range(1,16):
		sheettgt.cell(row=row,column=y).border = Border(left=bd, right=bd)

	tc_count=tc_count+1
	row=row+1
	sheettgt.cell(row=row,column=1).value =app
	sheettgt.cell(row=row,column=2).value =sub
	sheettgt.cell(row=row,column=3).value ="Verification of Count of records in "+targetTable+" table"
	sheettgt.cell(row=row,column=4).value ="COUNT_VALIDATION_"+targetTable+'_TC00'+str(tc_count)
	sheettgt.cell(row=row,column=5).value ="Complete"
	sheettgt.cell(row=row,column=6).value ="Step 1"
	sheettgt.cell(row=row,column=7).value ="Verify whether the count of records in "+targetTable+" table is matching aginst the count of records in Source table."
	sheettgt.cell(row=row,column=8).value ="Count of records in "+targetTable+" table should be matching aginst the count of records in Source table."
	sheettgt.cell(row=row,column=9).value ="Medium"
	sheettgt.cell(row=row,column=10).value ='No'
	sheettgt.cell(row=row,column=11).value ="Automated"
	sheettgt.cell(row=row,column=12).value ="Workflow"
	sheettgt.cell(row=row,column=13).value ="Manual"
	sheettgt.cell(row=row,column=14).value =tcDetails["regcat"]
	sheettgt.cell(row=row,column=15).value =tcDetails["regstat"]
	for y in range(1,16):
		sheettgt.cell(row=row,column=y).border = Border(left=bd, right=bd)

	tc_count=tc_count+1
	row=row+1
	sheettgt.cell(row=row,column=1).value =app
	sheettgt.cell(row=row,column=2).value =sub
	sheettgt.cell(row=row,column=3).value ="Verification of Structure of "+targetTable+" table"
	sheettgt.cell(row=row,column=4).value ="DDL_VALIDATION_"+targetTable+'_TC00'+str(tc_count)
	sheettgt.cell(row=row,column=5).value ="Complete"
	sheettgt.cell(row=row,column=6).value ="Step 1"
	sheettgt.cell(row=row,column=7).value ="Verify whether the structure of "+targetTable+" table is as per the PDM ( Physical Data Model)."
	sheettgt.cell(row=row,column=8).value ="The structure of "+targetTable+" table should be as per the PDM ( Physical Data Model)."
	sheettgt.cell(row=row,column=9).value ="Medium"
	sheettgt.cell(row=row,column=10).value ='No'
	sheettgt.cell(row=row,column=11).value ="Automated"
	sheettgt.cell(row=row,column=12).value ="Workflow"
	sheettgt.cell(row=row,column=13).value ="Manual"
	sheettgt.cell(row=row,column=14).value =tcDetails["regcat"]
	sheettgt.cell(row=row,column=15).value =tcDetails["regstat"]
	for y in range(1,16):
		sheettgt.cell(row=row,column=y).border = Border(left=bd, right=bd, bottom=bd)

	#sheettgt.set_column('A:D',50)
	sheettgt.column_dimensions['A'].width = 30
	sheettgt.column_dimensions['B'].width = 70
	sheettgt.column_dimensions['C'].width = 50
	sheettgt.column_dimensions['D'].width = 50
	sheettgt.column_dimensions['E'].width = 20
	sheettgt.column_dimensions['F'].width = 15
	sheettgt.column_dimensions['G'].width = 150
	sheettgt.column_dimensions['H'].width = 90
	sheettgt.column_dimensions['I'].width = 15
	sheettgt.column_dimensions['J'].width = 25
	sheettgt.column_dimensions['K'].width = 25
	sheettgt.column_dimensions['L'].width = 20
	sheettgt.column_dimensions['M'].width = 10
	sheettgt.column_dimensions['N'].width = 25
	sheettgt.column_dimensions['O'].width = 25

	wbtgt.save(desfile)

	print "\t\t\tCOMPLETED "
	return targetTable

def ddl(ddlfilepath,dbuser,dbpass,dbhost,dbport,dbserv):
	print ddlfilepath,',',dbuser,',',dbpass,',',dbhost,',',dbport,',',dbserv

def spec2q(self):
	print 'SPEQ'
	
def main():

	root = tk.Tk()
	root.title('SPARTA - Standard Platform for AcceleRated Test Automation')
	page=Index(root)
	root.wm_geometry("700x400+300+100")
	root.resizable(0,0)
	root.mainloop()

if __name__ == '__main__':
	multiprocessing.freeze_support()
	main()